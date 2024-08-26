import json
import os
import subprocess
import time
import tkinter as tk
from logging import error
from tkinter import filedialog
from tkinter.messagebox import askyesno, showwarning, showinfo
from datetime import datetime
import openpyxl
from kivy import Config
from kivy.app import App
from kivy.core.window import Window
from kivy.properties import ObjectProperty
from kivy.lang import Builder
from kivy.uix.screenmanager import ScreenManager, Screen
from Scripts.addingUsers import demoData
from Scripts.batches import batches
from Scripts.categoryGroups import categoryGroups
from Scripts.dataLibrary import dataLibrary
from Scripts.enableInsights import sEditorVisible, newEnableInsights
from Scripts.expressionsFormats import expressionsFormats
from Scripts.mergeCategories import mergeCategories
from Scripts.newFactField import newFactField
from Scripts.newProject import newProject
from Scripts.payload import payload
from Scripts.thFactor import thFactor
from Scripts.postMan import postManRequests
from Scripts.toolBoox import verify_json
from Scripts.toolBoox.toolBoox import rewriteText, verifyPath, openKB, getFile, readJson, getPath, getSolution, \
	getChannels
from Scripts.updateFactAttribute import updateFactAttribute
from Scripts.updateFactCategory import updateFactCategory

Builder.load_file('Scripts/Source/alerts.kv')
fileManger = 'Scripts/Source/fileManger.json'
root = tk.Tk()
root.withdraw()
root.destroy()


class MenuWindow(Screen):
	solutionPath = ObjectProperty(None)

	# channels = list(os.listdir(getSolution(getPath('solution'))))

	def openFile(self, name):
		if name == "expressionsFormats":
			showinfo("Attention", "Please note the following steps: \n\t* you must \"en\" col even if it have the same data. \n\t* you must put \' for the amount, like this : \n\t\'###,###,###.00\'")
		os.startfile(os.path.normpath(getFile(name + "_raw")), 'edit')

	def selected(self, name, filterX):
		checkPath = filedialog.askdirectory(initialdir=os.path.normpath(SettingsWindow().currentDefaultPath))
		if checkPath == "":
			if askyesno('Confirmation', 'Please choose a project'):
				self.selected(name, filterX)
			else:
				return
		verifyPath(name, filterX, checkPath)
		operations = os.path.join(getSolution(getPath('solution'), True), 'pack_dscr.properties')
		if os.path.exists(operations):
			mvnTime = ""
			newTime = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(os.path.getmtime(checkPath)))
			with open(operations, "r+") as f:
				d = f.readlines()
				for prop in d:
					if 'PACKAGING_TIMESTAMP' in prop:
						raw = prop.split('=')[-1]
						mvnTime = raw.replace('T', ' ').replace('Z', '')
			newTimeStamp = datetime.strptime(newTime, '%Y-%m-%d %H:%M:%S').timestamp()
			mvnTimeStamp = datetime.strptime(mvnTime[:-1], '%Y-%m-%d %H:%M:%S').timestamp()
			if mvnTimeStamp <= newTimeStamp:
				return

		if askyesno('Confirmation', 'Do you need to install MAVENs?'):
			p = subprocess.run('mvn clean install -DskipTests -U', shell=True, cwd=checkPath)
			if p.returncode != 0:
				error('Maven: couldn\'t complete the build for ' + checkPath)

	def openKB(self, root):
		openKB(root.name)

	def cleanExcels(self):
		files = readJson(fileManger)
		for file in files.keys():
			if file == 'expressionsFormats_raw':
				df = openpyxl.load_workbook(files[file])
				sheet = df['Sheet1']
				sheet.delete_cols(2, sheet.max_column)
				df.save(files[file])
			elif '_raw' in file:
				df = openpyxl.load_workbook(files[file])
				sheet = df['Sheet1']
				sheet.delete_rows(2, sheet.max_row)
				df.save(files[
							file])  # df = df.head(1)  # df.to_excel(files[file], index=False)  # fileName = files[file]  # book = openpyxl.load_workbook(files[file])  #  # lines = False  # for row in sheet.rows:  #     if lines:  #         for cell in row:  #             cell.value = None  #     else:  #         lines = not lines  # book.save(fileName)  # raise SystemExit


class enableInsightsWindow(Screen):
	Builder.load_file('Scripts/enableInsights/enableInsights.kv')

	def runFunc(self):
		# transfer.main([self.name, 'active'])
		newEnableInsights.main([getFile(self.name)])

	pass


class categoryGroupsWindow(Screen):
	Builder.load_file('Scripts/categoryGroups/categoryGroups.kv')

	def checkBoxClick(self, instance):
		self.ids.checkBoxs.text = instance

	def runFunc(self):
		canRun = False
		num_of_lang = self.ids.langNum.text
		languages = [self.ids.language1.text, self.ids.language2.text, self.ids.language3.text, self.ids.language4.text]
		if self.ids.checkBoxs.text != "" and languages.count("") != 4:
			canRun = True

		if canRun:
			categoryGroups.main([getFile(self.name + "_raw"), num_of_lang, languages, self.ids.checkBoxs.text])
		else:
			print("Please fill up relevant values")

	pass


class newFieldToFactWindow(Screen):
	Builder.load_file('Scripts/newFactField/newFactField.kv')

	def runFunc(self):
		newFactField.main([self.fact.text, self.field.text, self.value.text, self.type.text])

	pass


class updateFactAttributeWindow(Screen):
	Builder.load_file('Scripts/updateFactAttribute/updateFactAttribute.kv')

	def runFunc(self):
		updateFactAttribute.main([self.fact.text, self.field.text, self.value.text, self.oldVal.text])

	pass


class mergeCategoriesWindow(Screen):
	Builder.load_file('Scripts/mergeCategories/mergeCategories.kv')

	def runFunc(self):
		bUsers = self.ids.BUsers.active
		JInsightFacts = self.ids.Insights.active
		mergeCategories.main([getFile(self.name + "_raw"), bUsers, JInsightFacts])


class updateFactCategoryWindow(Screen):
	Builder.load_file('Scripts/updateFactCategory/updateFactCategory.kv')

	def runFunc(self):
		updateFactCategory.main([self.fact.text, self.field.text, self.value.text, self.oldVal.text])

	pass


class batchesWindow(Screen):
	Builder.load_file('Scripts/batches/batches.kv')
	corePath = ObjectProperty(None)
	solutionPath = ObjectProperty(None)
	checkBoxs = ObjectProperty(None)

	def methodType(self, methodType):
		if methodType != "Choose a method":
			self.ids.batch.disabled = False
		if methodType == "Choose a method" and self.ids.channel.text == "":
			self.ids.batch.disabled = True
			self.ids.method.text = methodType

	def batchType(self, batchType):
		flows = {"data-assets": "data-assets-flow", "push-notification-only-send": "push-notification-only-send-flow",
				 "act-data-assets": "data-assets-microsavings-eligibility-flow", "purging": "purging-flow"}
		try:
			self.ids.channel.text = getChannels(getSolution(App.get_running_app().root.ids.Menu_Window.ids.solutionPath.text), True)
		except Exception as e:
			error('Path Error:' + e.__str__()[e.__str__().index(']') + 1:])
			showwarning("No solution", "Please choose a solution before configuring the batch.")
			self.ids.batch.text = "Choose a batch"

		if self.ids.channel.text != "":
			self.ids.flowId.text = flows[batchType]
			if batchType != "Choose a batch":
				self.ids.startingTime.disabled = False
				self.ids.endingTime.disabled = False
				self.ids.excludeDays.disabled = False
				self.ids.intervalInMin.disabled = False
				self.ids.flowId.disabled = False
				self.ids.maxResult.disabled = False
				self.ids.importPagingSize.disabled = False
				self.ids.threadSplitter.disabled = False
				self.ids.recurrencePattern.disabled = False
				self.ids.taskType.disabled = False
				self.ids.groupName.disabled = False
				self.ids.channel.disabled = False
				self.ids.autoGenerate.disabled = False
				if batchType != "purging":
					self.ids.importMethods.disabled = False
					self.ids.pushNotificationOptIn.disabled = False
					self.ids.serverSynchronization.disabled = True
					self.ids.Context.disabled = True

					if self.ids.importMethods.text == "autoRegister":
						self.ids.API.disabled = False
						self.ids.Context.disabled = False

				elif batchType == "purging":
					self.ids.serverSynchronization.disabled = False
					self.ids.Context.disabled = False
					self.ids.API.disabled = True
					self.ids.importMethods.disabled = True
					self.ids.pushNotificationOptIn.disabled = True
					self.ids.Context.text = "insightsDefault"
					self.ids.taskType.text = "etlBatchContinues"
		pass

	def methods_update(self, value):
		print(self)
		if value == "autoRegister":
			self.ids.API.disabled = False
			self.ids.Context.disabled = False
		else:
			self.ids.API.disabled = True
			self.ids.Context.disabled = True

		pass

	pass

	def runFunc(self):
		batchData = {"method": self.ids.method.text, "id": self.ids.batch.text, "active": True,
					 "channel": self.ids.channel.text, "QA": self.ids.QaBatch.active,
					 "startingTime": self.ids.startingTime.text, "endingTime": self.ids.endingTime.text,
					 "excludeDays": self.ids.excludeDays.text, "recurrencePattern": self.ids.recurrencePattern.text,
					 "flowId": self.ids.flowId.text, "maxResult": int(self.ids.maxResult.text),
					 "importPagingSize": int(self.ids.importPagingSize.text),
					 "threadSplitter": int(self.ids.threadSplitter.text),
					 "importMethods": [self.ids.importMethods.text], "taskType": self.ids.taskType.text,
					 "intervalInMin": int(self.ids.intervalInMin.text),
					 "pushNotificationOptIn": self.ids.pushNotificationOptIn.text, "groupName": self.ids.groupName.text,
					 "autoGenerate": self.ids.autoGenerate.text,
					 "serverSynchronization": self.ids.serverSynchronization.text, "API": self.ids.API.text,
					 "context": self.ids.Context.text}
		if batchData["pushNotificationOptIn"] == "True":
			batchData["pushNotificationOptIn"] = True
		else:
			batchData["pushNotificationOptIn"] = False
		batches.main([batchData])


class demoDataWindow(Screen):
	Builder.load_file('Scripts/addingUsers/demoData.kv')

	def runFunc(self):
		bUsers = self.ids.BUsers.active
		modified = self.ids.modified.active
		properties = {"LocalCurrency": self.ids.LocalCurrency.text, "ForeignCurrency": self.ids.ForeignCurrency.text,
					  "CountryName": self.ids.CountryName.text, "CountryCode": self.ids.CountryCode.text,
					  "Factor": int(self.ids.Factor.text)}
		demoData.main([properties, bUsers, modified, getFile(self.name + "_raw")])

	pass


class dataLibraryWindow(Screen):
	Builder.load_file('Scripts/dataLibrary/dataLibrary.kv')
	solutionPath = ObjectProperty(None)

	def checkBoxClick(self, instance):
		print(self.ids.deactivated.active)

	def runFunc(self, file):
		excel = getFile(file + "_raw")
		dataLibrary.main([excel, self.ids.dataLibrary.active, self.ids.deactivated.active])

	pass


class sEditorVisibleWindow(Screen):
	Builder.load_file('Scripts/enableInsights/sEditorVisible.kv')

	def runFunc(self):
		sEditorVisible.main([getFile(self.name + "_raw")])

	pass


class expressionsFormatsWindow(Screen):
	Builder.load_file('Scripts/expressionsFormats/expressionsFormats.kv')

	def runFunc(self):
		expressionsFormats.main([getFile(self.name + "_raw")])

	pass


class breakingChangeWindow(Screen):
	Builder.load_file('Scripts/breakingChange/breakingChange.kv')

	# def runFunc(self, path):
	#     if valPath(path):
	#         status = sEditorVisible.main([path, self.name])
	#         print(status)

	pass


class newProjectWindow(Screen):
	Builder.load_file('Scripts/newProject/newProject.kv')

	def checkBoxClick(self, check):
		pass

	def runFunc(self):
		newProject.main([self.ids.channel.text])

	pass


class payloadWindow(Screen):
	Builder.load_file('Scripts/payload/payload.kv')
	payloadJson = str()

	def getPayload(self, jsonText: str(), popUpSelf):
		if verify_json(jsonText):
			payloadWindow.payloadJson = jsonText
			popUpSelf.dismiss()
		else:
			showwarning("Json verification", "The JSON You entered is not validated. \nPlease check the JSON")
		pass

	def dataAssetsUrl(self, value: bool):
		self.ids.RemoteAssetsUrl.disabled = not value

		pass

	def runFunc(self):
		if self.ids.widgetType.text == "Selecet widget type":
			print("NO GO")
		else:
			payload.main([self.ids.widgetType.text, self.ids.Lang.text, self.ids.RemoteAssetsActivation.active,
						  self.ids.RemoteAssetsUrl.text, self.ids.darkMode.active, payloadWindow.payloadJson])

	pass


class PostManWindow(Screen):
	Builder.load_file('Scripts/postMan/postMan.kv')

	# solutionPath = ObjectProperty(None)
	# ipAddress = ObjectProperty(None)
	# channel = ObjectProperty(None)
	# checkBoxs = ObjectProperty(None)
	# user = ObjectProperty(None)
	# context = ObjectProperty(None)
	#
	def checkBoxDate(self, date):
		if date in ('01/03/2017', '12/22/2016'):
			self.ids.checkBoxs.text = date
		else:
			self.ids.date.disabled = False

	def submit(self, date):
		self.ids.checkBoxs.text = date

	def checkDate(self, date):
		dateFormat = '%m/%d/%Y'
		try:
			datetime.datetime.strptime(date, dateFormat)
		except ValueError:
			print("Incorrect data format, should be MM-DD-YYYY")

	def requestProperties(self):
		RequestsWindow.ipAddress = self.ids.ipAddress.text
		RequestsWindow.channel = self.ids.channel.text
		RequestsWindow.checkBoxs = self.ids.checkBoxs.text
		RequestsWindow.user = self.ids.user.text
		RequestsWindow.context = self.ids.context.text


class RequestsWindow(Screen):
	Builder.load_file('Scripts/postMan/requests.kv')
	ipAddress = ''
	channel = ''
	checkBoxs = ''
	user = ''
	context = ''

	def runFunc(self, api):
		postManRequests.main([RequestsWindow.ipAddress, RequestsWindow.channel, RequestsWindow.checkBoxs, api,
							  RequestsWindow.user, RequestsWindow.context])

	pass


class ThFactorWindow(Screen):
	Builder.load_file('Scripts/thFactor/thFactor.kv')

	def runFunc(self):
		factor = float(self.ids.factor.text)
		thFactor.main([factor])

	pass


class jsonsWindow(Screen):
	def runFunc(self):
		updateFactCategory.main([])

	pass


class SettingsWindow(Screen):
	if not os.path.exists(getFile('settings')):
		print("check")
		pathRootJson = {"pathRoot": "", "intelliJRoot": "", "corePath": "", "modelPath": "", "EBPath": ""}
		f = open(getFile('settings'), "x")
		f.write(json.dumps(pathRootJson, indent=4))
		f.close()
	currentDefaultPath = readJson(getFile('settings'))['pathRoot']
	currentIntelliJPath = readJson(getFile('settings'))['intelliJRoot']
	currentCorePath = readJson(getFile('settings'))['corePath']
	currentModelPath = readJson(getFile('settings'))['modelPath']
	currentEBPath = readJson(getFile('settings'))['EBPath']

	def dpath(self, path, path_filter):
		rewriteText(getFile(self.name), path, path_filter)

	pass


class WindowManger(ScreenManager):
	Window.size = (1000, 800)
	Window.minimum_height = 755
	Window.minimum_width = 800
	# #
	pass


class ImplementationApp(App):
	def build(self):
		self.icon = "persoLogo.png"
		return Builder.load_file('main.kv')
		Config.set('graphics', 'width', '1000')
		Config.set('graphics', 'height', '1000')
		Config.set('graphics', 'minimum_width', '800')
		Config.set('graphics', 'minimum_height', '755')


if __name__ == '__main__':
	ImplementationApp().run()
